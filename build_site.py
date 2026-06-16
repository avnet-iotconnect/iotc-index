#!/usr/bin/env python3
"""
build_site.py — reads iotc-index-catalog.xlsx and produces:
  - index.json   (what the page renders)
  - AUDIT.md     (gap report: missing boards, orphan boards, image gaps,
                  uncatalogued org repos, incomplete listings)

Run locally:           python build_site.py
In CI with live facts:  GITHUB_TOKEN=xxx python build_site.py
"""
import os, re, json, sys, urllib.request
from openpyxl import load_workbook

XLSX = os.environ.get("CATALOG_XLSX", "iotc-index-catalog.xlsx")
OUT  = os.environ.get("OUT_DIR", ".")
TOKEN = os.environ.get("GITHUB_TOKEN", "")

def slug(s): return re.sub(r'[^a-z0-9]+','-',str(s).lower()).strip('-') or "x"
def split(s): return [x.strip() for x in str(s or "").replace(";",",").split(",") if x.strip()]

wb = load_workbook(XLSX, data_only=True)
cfg = {r[0]: r[1] for r in wb["Config"].iter_rows(min_row=2, values_only=True) if r and r[0]}
ORG = os.environ.get("ORG", cfg.get("ORG","avnet-iotconnect"))
IMAGE_BASE = cfg.get("IMAGE_BASE","")
DEFAULT_STATUS = cfg.get("DEFAULT_STATUS","beta")

def resolve_image(v):
    if not v: return None
    s=str(v).strip()
    if re.match(r'^(https?:|data:|//)', s): return s
    return (IMAGE_BASE or "") + s

# ---------- Boards ----------
brows = list(wb["Boards"].iter_rows(min_row=2, values_only=True))
bhdr  = [c.value for c in wb["Boards"][1]]
def bcol(row,name): 
    i=bhdr.index(name); return row[i] if i<len(row) else None
boards_by_pn = {}; boards_by_name = {}; board_defs = {}
for row in brows:
    if not row or not (bcol(row,"Board Name") or bcol(row,"Part Number")): continue
    if str(bcol(row,"Include") or "yes").lower()=="no": continue
    pn = str(bcol(row,"Part Number") or "").strip()
    nm = str(bcol(row,"Board Name") or "").strip()
    sg = slug(pn or nm)
    quals=[]
    if str(bcol(row,"AWS Qualified") or "").lower()=="yes": quals.append("AWS Qualified")
    if str(bcol(row,"Greengrass") or "").lower()=="yes": quals.append("Greengrass")
    d = {"slug":sg,"vendor":str(bcol(row,"Manufacturer") or "Other").strip(),"name":nm or pn,
         "partNumber":pn,"image":resolve_image(bcol(row,"Image File")),"link":bcol(row,"Product Link"),
         "qualifications":quals,"tags":split(bcol(row,"Tags"))}
    board_defs[sg]=d
    if pn: boards_by_pn[pn.lower()]=sg
    if nm: boards_by_name[nm.lower()]=sg

# ---------- Listings ----------
lrows = list(wb["Listings"].iter_rows(min_row=2, values_only=True))
lhdr  = [c.value for c in wb["Listings"][1]]
def lcol(row,name):
    i=lhdr.index(name); return row[i] if i<len(row) else None

missing_boards = {}   # ref -> set(listing names)
def resolve_ref(ref, listing_name):
    k=ref.lower()
    if k in boards_by_pn: return boards_by_pn[k]
    if k in boards_by_name: return boards_by_name[k]
    missing_boards.setdefault(ref, set()).add(listing_name)
    sg="x-"+slug(ref)
    board_defs.setdefault(sg, {"slug":sg,"vendor":"Other","name":ref,"partNumber":"","image":None,"qualifications":[],"tags":[]})
    return sg

listings=[]
for row in lrows:
    if not row or not lcol(row,"Name"): continue
    if str(lcol(row,"Include") or "yes").lower()=="no": continue
    name=str(lcol(row,"Name")).strip()
    refs=[resolve_ref(r, name) for r in split(lcol(row,"Boards"))]
    listings.append({
        "name":name, "repo":(str(lcol(row,"Repo")).strip() or None) if lcol(row,"Repo") else None,
        "category":str(lcol(row,"Type") or "uncategorized").strip().lower(),
        "status":str(lcol(row,"Status") or DEFAULT_STATUS).strip().lower(),
        "languages":split(lcol(row,"Languages")), "features":split(lcol(row,"Topics")),
        "boards":refs, "description":str(lcol(row,"Description") or "").strip(),
        "url":(str(lcol(row,"Link")).strip() or None) if lcol(row,"Link") else None,
    })

# ---------- live GitHub facts ----------
def gh_get(path):
    req=urllib.request.Request("https://api.github.com"+path,
        headers={"Accept":"application/vnd.github+json","User-Agent":"iotc-index",
                 **({"Authorization":f"Bearer {TOKEN}"} if TOKEN else {})})
    with urllib.request.urlopen(req, timeout=30) as r: return json.load(r)
facts={}; org_repos=[]
try:
    page=1
    while True:
        batch=gh_get(f"/orgs/{ORG}/repos?per_page=100&page={page}&type=public&sort=updated")
        org_repos+=batch
        if len(batch)<100: break
        page+=1
    for r in org_repos:
        facts[r["name"]]={"description":r.get("description") or "","languages":[r["language"]] if r.get("language") else [],
                          "stars":r.get("stargazers_count",0),"updated":r.get("updated_at"),"archived":r.get("archived")}
except Exception as e:
    print(f"[info] GitHub facts unavailable ({e}); using workbook data only.", file=sys.stderr)

# ---------- assemble index ----------
def board_obj(sg):
    b=board_defs[sg]; o={"slug":sg,"vendor":b["vendor"],"name":b["name"]}
    if b.get("partNumber"): o["partNumber"]=b["partNumber"]
    if b.get("image"): o["image"]=b["image"]
    return o

out=[]; used_repos=set()
for L in listings:
    live=facts.get(L["repo"]) if L["repo"] else None
    if L["repo"]: used_repos.add(L["repo"])
    boards=[board_obj(s) for s in L["boards"]]
    base={"repo":L["repo"],"url":L["url"] or (f'https://github.com/{ORG}/{L["repo"]}' if L["repo"] else None),
          "displayName":L["name"],"description":L["description"] or (live or {}).get("description",""),
          "category":L["category"],"status":L["status"],
          "languages":L["languages"] or (live or {}).get("languages",[]),"features":L["features"],
          "stars":(live or {}).get("stars",0),"updated":(live or {}).get("updated","2026-06-01T00:00:00Z"),"hidden":False}
    if L["category"]=="sample" and len(boards)>1:
        for b in boards: out.append({**base,"id":f'{slug(L["name"])}::{b["slug"]}',"boards":[b],"board":b,"manufacturers":[b["vendor"]]})
    else:
        out.append({**base,"id":slug(L["name"]),"boards":boards,"board":boards[0] if boards else None,
                    "manufacturers":sorted({b["vendor"] for b in boards}) if boards else []})

vis=[r for r in out if not r["hidden"]]
mfrs=sorted({m for r in vis for m in r["manufacturers"] if m and m!="Other"})
busd=sorted({b["slug"]:b for r in vis for b in r["boards"]}.values(), key=lambda b:b["name"])
index={"org":ORG,"generated":__import__("datetime").datetime.now(__import__("datetime").timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),"imageBase":IMAGE_BASE,
       "facets":{"manufacturers":mfrs,"boards":busd},
       "counts":{"total":len(vis),"manufacturers":len(mfrs),"boards":len(busd),
                 "sdks":sum(1 for r in vis if r["category"] in("sdk","library")),
                 "examples":sum(1 for r in vis if r["category"]=="sample")},
       "repos":out}
os.makedirs(OUT, exist_ok=True)
json.dump(index, open(os.path.join(OUT,"index.json"),"w"), indent=2)

# ---------- AUDIT ----------
used_board_slugs={b["slug"] for r in vis for b in r["boards"]}
orphan=[d for s,d in board_defs.items() if not s.startswith("x-") and s not in used_board_slugs]
no_image=[d for s,d in board_defs.items() if not s.startswith("x-") and not d.get("image")]
uncatalogued=[r for r in org_repos if not r.get("archived") and r["name"] not in used_repos] if org_repos else []
incomplete=[L["name"] for L in listings if (not L["features"]) or (not L["description"] and not L["repo"])]

A=["# /IOTCONNECT Index — Audit Report","",
   f"_Generated {index['generated']}_  ·  {len(vis)} listings · {len(busd)} boards in use · {len(mfrs)} manufacturers","",
   "## ⚠ Boards referenced but missing from the Boards sheet"]
if missing_boards:
    for ref, who in sorted(missing_boards.items()):
        A.append(f"- **{ref}** — referenced by: {', '.join(sorted(who))}")
else: A.append("- none — every referenced board is defined.")
A+=["","## Boards with no image"]
A += [f"- {d['vendor']} · {d['name']} ({d.get('partNumber') or 'no PN'})" for d in no_image] or ["- none."]
A+=["","## Boards not used by any listing (orphans)"]
A += [f"- {d['vendor']} · {d['name']} ({d.get('partNumber') or 'no PN'})" for d in orphan] or ["- none."]
A+=["","## Org repos with no listing (candidates to add)"]
if org_repos:
    A += [f"- {r['name']} — {r.get('description') or 'no description'}" for r in uncatalogued] or ["- none — every public repo has a listing."]
else:
    A.append("- (run with GITHUB_TOKEN to detect uncatalogued repos)")
A+=["","## Listings missing description or topics"]
A += [f"- {n}" for n in incomplete] or ["- none."]
open(os.path.join(OUT,"AUDIT.md"),"w").write("\n".join(A)+"\n")

print(f"index.json: {len(vis)} listings, {len(mfrs)} manufacturers, {len(busd)} boards")
print(f"AUDIT: {len(missing_boards)} missing-board refs, {len(no_image)} no-image, {len(orphan)} orphan, "
      f"{len(uncatalogued)} uncatalogued repos, {len(incomplete)} incomplete listings")
