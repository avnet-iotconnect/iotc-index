#!/usr/bin/env python3
"""
build_catalog.py — ONE-TIME migration / enrichment generator (PROVENANCE ONLY).

⚠ DO NOT re-run against an edited workbook. This regenerates
`../iotc-index-catalog.xlsx` from the *original* pre-migration snapshot
(_wb_dump.json) plus the harvested overlay, and would discard any later
hand-edits made to the workbook. The workbook itself is now the source of truth;
edit it and run ../build_site.py. This file is kept only to document how the
enriched catalog was produced.

Fuses the original workbook data (_wb_dump.json) with the harvested overlay
(_overlay.json) and the partner-image map (_partner_img_map.json) into the
enriched source workbook:

  - Listings : existing + newly catalogued repos, with topics filled in
  - Boards   : existing + reconciled boards, broken refs repaired,
               new "Image Local" column (local fallback under assets/boards/)
  - Resources: NEW sheet — per-board Buy / QuickStart / Developer / Demo /
               Webinar / Video / Blog / Doc links harvested from the partner READMEs
  - Config / Instructions : refreshed
"""
import os, json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Inputs live next to this script (migration/); the workbook is written to the repo root.
HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
wb_dump   = json.load(open(os.path.join(HERE, "_wb_dump.json"), encoding="utf-8"))
overlay   = json.load(open(os.path.join(HERE, "_overlay.json"), encoding="utf-8"))
img_map   = json.load(open(os.path.join(HERE, "_partner_img_map.json"), encoding="utf-8"))

def rows_to_dicts(sheet):
    hdr = sheet[0]
    return [dict(zip(hdr, r + [""] * (len(hdr) - len(r)))) for r in sheet[1:]]

L_in = rows_to_dicts(wb_dump["Listings"])
B_in = rows_to_dicts(wb_dump["Boards"])

# ----------------------------------------------------------------------------
# 1) PartNumber -> local image filename (from partner README purchase sections)
# ----------------------------------------------------------------------------
def norm_pn(s): return re.sub(r"\s+", "", str(s or "")).lower()
local_img = {}
for e in img_map:
    pn = norm_pn(e.get("purchaseText"))
    img = (e.get("img") or "").strip()
    if pn and img:
        # filenames were normalised on disk: non-breaking hyphen -> '-', spaces -> '_'
        img = img.replace("‑", "-").replace(" ", "_")
        local_img.setdefault(pn, img)
# manual locals for boards whose Azure image is broken / has spaces
local_img.setdefault(norm_pn("MSC SM2S-IMX8PLUS"), "board_SM_MSC_IMX8PLUS_small.png")
# the ST partner README reused the N6 photo for the MP2 DK section — use the correct local file
local_img[norm_pn("STM32MP257F-DK")] = "STM32MP257F-DK.jpg"

# ----------------------------------------------------------------------------
# 2) Board-ref reconciliation (rename map + new boards)
# ----------------------------------------------------------------------------
rename = {}
for f in overlay["boardRefFixes"]:
    rename[f["badRef"].strip().lower()] = f["renameTo"].strip()

NEW_BOARD_IMG = {}  # added boards fall back to a vendor monogram placeholder

# Azure-hosted images that 404 -> blank so the card shows a clean monogram
DEAD_AZURE_IMG = {"MIC-710AIL"}
# Vendor product pages that 404 -> blank (so a no-image/no-resource orphan drops from the gallery)
DEAD_PRODUCT_LINK = {"MIC-710AIL"}

# Listing links that 404 (dead branch / moved file) or point at a PR -> repaired
LISTING_LINK_FIX = {
    "https://github.com/avnet-iotconnect/iotc-st-image-classification/tree/initial":
        "https://github.com/avnet-iotconnect/iotc-st-image-classification",
    "https://github.com/avnet-iotconnect/iotc-freertos-ck-ra6m5-v2-pmod/blob/master/QUICKSTART.md":
        "https://github.com/avnet-iotconnect/iotc-freertos-ck-ra6m5-v2-pmod",
    "https://github.com/Avnet/QCS6490-Vision-AI-Demo/pull/1":
        "https://github.com/Avnet/QCS6490-Vision-AI-Demo",
}

# Per-listing corrections (review findings): rename / retag / rebind / give url to dead-end demos
LISTING_PATCH = {
    "Edge AI Inference Pipeline (Jetson Orin NX)": {
        "Repo": "iotc-jetson-demo", "Link": "https://github.com/avnet-iotconnect/iotc-jetson-demo",
        "Languages": "Python", "Topics": "edge-ai, vision, npu"},
    "FPGA-Driven Industrial Gesture Recognition": {"Topics": "edge-ai, vision, gesture, fpga"},
    "Image Classification and Remote Training": {"Topics": "edge-ai, vision, mpu, ota"},
    "Edge AI Solutions featuring Jetson": {"Name": "Telehealth Mobile Gateway",
        "Topics": "ble, gateway, healthcare, telemetry"},
    "STM32N6 Edge AI Object Detection": {"Boards": "STM32N6570-DK"},  # DA16K Pmod is an accessory, drop the split card
    "STSAFE Authentification": {"Name": "STSAFE Secure Authentication",
        "Topics": "security, telemetry",
        "Description": "Hardware-anchored device authentication using the STSAFE-A secure element on the STM32U5. "
                       "The board provisions cryptographic credentials and connects to /IOTCONNECT with a verified "
                       "identity, demonstrating telemetry and command flows built on a hardware root of trust."},
}

# Drop listings made redundant by dedicated-repo listings added below/in the overlay
LISTING_DROP_NAMES = {"PSOC6 AI-Kit Human Activity", "Smart Home Baby Monitor"}

# Per-board primary buy-link overrides (fix wrong-product links)
BUY_OVERRIDE = {
    "KITPSE84AITOBO1": "https://www.newark.com/infineon/kitpse84aitobo1/ai-eval-kit-32bit-arm-cortex-m55f/dp/49AM4459",
    "KIT_PSE84_EVAL":  "https://www.newark.com/infineon/kitpse84evaltobo1/eval-kit-32bit-arm-cortex-m55f/dp/49AM4460",
}

# Completeness: flagship SDKs missing from the catalog (AUDIT gaps)
SUPPLEMENTAL_LISTINGS = [
    {"Name": "/IOTCONNECT Python SDK", "Type": "sdk", "Status": "stable", "Repo": "iotc-python-sdk",
     "Languages": "Python", "Topics": "telemetry, commands, mpu", "Boards": "",
     "Description": "Full-featured Python SDK for /IOTCONNECT — the most common starting point for desktop, "
                    "server and embedded Linux device-to-cloud integrations.",
     "Link": "https://github.com/avnet-iotconnect/iotc-python-sdk", "Include": "yes"},
    {"Name": "/IOTCONNECT Python REST API", "Type": "library", "Status": "stable", "Repo": "iotc-python-rest-api",
     "Languages": "Python", "Topics": "telemetry, commands", "Boards": "",
     "Description": "Python client for the /IOTCONNECT REST API — manage devices, templates and telemetry programmatically.",
     "Link": "https://github.com/avnet-iotconnect/iotc-python-rest-api", "Include": "yes"},
    {"Name": "/IOTCONNECT nRF Connect SDK", "Type": "sdk", "Status": "beta", "Repo": "iotc-nrf-sdk",
     "Languages": "C", "Topics": "cellular, ble, mcu", "Boards": "nRF6943, nRF52840-DK",
     "Description": "/IOTCONNECT SDK for Nordic nRF9160 and nRF52 devices — connect the Thingy:91 and nRF52840 DK "
                    "over cellular or BLE to the cloud.",
     "Link": "https://github.com/avnet-iotconnect/iotc-nrf-sdk", "Include": "yes"},
]

# ----------------------------------------------------------------------------
# 3) Build the Boards sheet (dedupe duplicate PNs, add Image Local)
# ----------------------------------------------------------------------------
BCOLS = ["Manufacturer", "Board Name", "Part Number", "Image File", "Image Local",
         "Product Link", "AWS Qualified", "Greengrass", "Tags", "Include"]

boards = {}      # pn(lower) -> row dict
def add_board(d):
    pn = str(d.get("Part Number", "")).strip()
    if not pn:
        return
    key = pn.lower()
    if key in boards:
        # merge: prefer rows that carry an image / qualifications
        cur = boards[key]
        for c in BCOLS:
            if not str(cur.get(c, "")).strip() and str(d.get(c, "")).strip():
                cur[c] = d[c]
        # keep an image if either row has one
        if not str(cur.get("Image File", "")).strip() and str(d.get("Image File", "")).strip():
            cur["Image File"] = d["Image File"]
        return
    boards[key] = {c: d.get(c, "") for c in BCOLS}

for b in B_in:
    row = {c: b.get(c, "") for c in BCOLS if c != "Image Local"}
    pn = str(b.get("Part Number", "")).strip()
    row["Image Local"] = local_img.get(norm_pn(pn), "")
    # fix the spaces-in-filename Azure image for MSC by blanking it (local takes over)
    if str(row.get("Image File", "")).strip() == "MSC SM2S-IMX8PLUS.png":
        row["Image File"] = ""
    if pn in DEAD_AZURE_IMG:
        row["Image File"] = ""
    if pn in DEAD_PRODUCT_LINK:
        row["Product Link"] = ""   # dead vendor page; lets the orphan card drop out of the gallery
    add_board(row)

for nb in overlay["newBoards"]:
    pn = nb["PartNumber"].strip()
    add_board({
        "Manufacturer": nb["Manufacturer"], "Board Name": nb["BoardName"],
        "Part Number": pn, "Image File": NEW_BOARD_IMG.get(pn, ""),
        "Image Local": local_img.get(norm_pn(pn), ""),
        "Product Link": nb.get("ProductLink", ""),
        "AWS Qualified": nb.get("AWSQualified", "no"),
        "Greengrass": nb.get("Greengrass", "no"),
        "Tags": nb.get("Tags", ""), "Include": "yes",
    })

valid_pns = {k for k in boards}

# ----------------------------------------------------------------------------
# 4) Listings: apply topics, repair board refs, enrich Sidewalk, add new repos
# ----------------------------------------------------------------------------
LCOLS = ["Name", "Type", "Status", "Repo", "Languages", "Topics", "Boards",
         "Description", "Link", "Include"]

topics_by_name = {t["listingName"].strip(): ", ".join(t["topics"]) for t in overlay["topicAssignments"]}

def fix_refs(boards_cell):
    out = []
    for r in re.split(r"[;,]", str(boards_cell or "")):
        r = r.strip()
        if not r:
            continue
        out.append(rename.get(r.lower(), r))
    # de-dup while preserving order
    seen, res = set(), []
    for x in out:
        if x.lower() not in seen:
            seen.add(x.lower()); res.append(x)
    return ", ".join(res)

listings = []
for l in L_in:
    row = {c: l.get(c, "") for c in LCOLS}
    name = str(row["Name"]).strip()
    if name in LISTING_DROP_NAMES:
        continue
    if not str(row.get("Topics", "")).strip() and name in topics_by_name:
        row["Topics"] = topics_by_name[name]
    row["Boards"] = fix_refs(row.get("Boards", ""))
    row["Link"] = LISTING_LINK_FIX.get(str(row.get("Link", "")).strip(), row.get("Link", ""))
    if name in LISTING_PATCH:
        row.update(LISTING_PATCH[name])
        row["Boards"] = fix_refs(row.get("Boards", ""))
    # enrich the repo-less "Sidewalk" listing with its real repo
    if name == "Sidewalk" and not str(row.get("Repo", "")).strip():
        row["Repo"] = "iotc-stm32-sidewalk"
        if not str(row.get("Topics", "")).strip():
            row["Topics"] = "lora, ble, telemetry"
    listings.append(row)

existing_repos = {str(r.get("Repo", "")).strip().lower() for r in listings if str(r.get("Repo", "")).strip()}
SKIP_NEW = {"iotc-stm32-sidewalk"}  # already merged into the existing "Sidewalk" listing

added = 0
for nl in overlay["newListings"]:
    repo = nl["repo"].strip()
    if repo.lower() in SKIP_NEW:
        continue
    listings.append({
        "Name": nl["name"].strip(), "Type": nl["type"].strip(), "Status": "beta",
        "Repo": repo, "Languages": nl.get("languages", ""), "Topics": nl.get("topics", ""),
        "Boards": fix_refs(nl.get("boards", "")), "Description": nl.get("description", ""),
        "Link": nl.get("link", "") or f"https://github.com/avnet-iotconnect/{repo}",
        "Include": "yes",
    })
    added += 1

for sl in SUPPLEMENTAL_LISTINGS:
    if sl["Repo"].lower() in existing_repos:
        continue
    row = {c: sl.get(c, "") for c in LCOLS}
    row["Boards"] = fix_refs(row.get("Boards", ""))
    listings.append(row)
    added += 1

# ----------------------------------------------------------------------------
# 5) Resources sheet (apply the 5 verified source-link fixes; dedupe)
# ----------------------------------------------------------------------------
URL_REPLACE = {
    # garbled trailing "Each" on a youtu.be id
    "https://youtu.be/mSETqaqMejQEach": "https://youtu.be/mSETqaqMejQ",
    # preprod/staging host -> public store
    "https://onesite-preprod.avnet.com/shop/us/products/renesas-electronics/da16200mod-devkt-3074457345643529171/":
        "https://www.avnet.com/shop/us/products/renesas-electronics/da16200mod-devkt-3074457345643529171/",
}
# drop these (wrong board / wrong SKU buy links, board ProductLink covers them)
DROP_URLS = {
    "https://www.avnet.com/shop/us/products/microchip/ev36w50a-3074457345653385127/",  # on AES-MC-SBC-IMXRT1176-G
    "https://www.avnet.com/shop/us/products/renesas-electronics/da16600mod-devkt-3074457345645205670/",  # on US159-DA16600EVZ
}

KIND_ORDER = {k: i for i, k in enumerate(
    ["buy", "quickstart", "developer", "demo", "webinar", "video", "blog", "doc", "info"])}

res_rows = []
seen_res = set()
mfr_info = {}   # manufacturer -> info url (collected, emitted once)
for r in overlay["resources"]:
    url = URL_REPLACE.get(r["url"].strip(), r["url"].strip())
    if not url or url in DROP_URLS:
        continue
    pn = r.get("boardPartNumber", "").strip()
    kind = r["kind"].strip()
    title = re.sub(r"&amp;", "&", r.get("title", "")).strip()
    mfr = r.get("manufacturer", "").strip()
    # fold manufacturer "More Information" links up to manufacturer-level (dedupe)
    if kind == "info" and "avnet.com/iotconnect/" in url:
        mfr_info.setdefault(mfr, url)
        continue
    ref = pn if pn else mfr
    reftype = "board" if pn else "manufacturer"
    key = (ref.lower(), kind, url.lower())
    if key in seen_res:
        continue
    seen_res.add(key)
    res_rows.append({"Ref": ref, "RefType": reftype, "Kind": kind, "Title": title, "URL": url})

# ensure every board with a known partner buy/product link has at least a buy resource
buy_by_pn = {row["Ref"].lower() for row in res_rows if row["Kind"] == "buy"}
for e in img_map:
    pn = (e.get("purchaseText") or "").strip()
    purl = (e.get("purchaseUrl") or "").strip()
    if not pn or not purl or pn.lower() in buy_by_pn:
        continue
    if pn.lower() in valid_pns and purl not in DROP_URLS:
        purl = URL_REPLACE.get(purl, purl)
        res_rows.append({"Ref": pn, "RefType": "board", "Kind": "buy",
                         "Title": f"Buy {pn}", "URL": purl})
        buy_by_pn.add(pn.lower())

# board-specific buy-link overrides (replace any harvested buy rows for these PNs)
for pn, url in BUY_OVERRIDE.items():
    res_rows = [r for r in res_rows if not (r["Kind"] == "buy" and r["Ref"].lower() == pn.lower())]
    res_rows.append({"Ref": pn, "RefType": "board", "Kind": "buy", "Title": f"Buy {pn} on Newark", "URL": url})

# manufacturer-level info rows
for mfr, url in sorted(mfr_info.items()):
    res_rows.append({"Ref": mfr, "RefType": "manufacturer", "Kind": "info",
                     "Title": f"{mfr} on /IOTCONNECT", "URL": url})

# sort: board resources grouped, by kind order
res_rows.sort(key=lambda r: (r["RefType"] != "board", r["Ref"].lower(), KIND_ORDER.get(r["Kind"], 99)))

# validate resource board refs
orphan_res = sorted({r["Ref"] for r in res_rows if r["RefType"] == "board" and r["Ref"].lower() not in valid_pns})

# ----------------------------------------------------------------------------
# 6) Write the workbook
# ----------------------------------------------------------------------------
WB = Workbook()
HFILL = PatternFill("solid", fgColor="0F1B2D")
HFONT = Font(bold=True, color="FFFFFF")

def write_sheet(name, cols, rows, widths=None):
    ws = WB.create_sheet(name)
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c); cell.fill = HFILL; cell.font = HFONT
        cell.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
    ws.freeze_panes = "A2"
    if widths:
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return ws

WB.remove(WB.active)

# Instructions
ins = WB.create_sheet("Instructions")
INSTRUCTIONS = [
    "/IOTCONNECT  Hardware & SDK Index — Source Workbook",
    "",
    "This workbook is the single source of truth for the public index at the Pages URL in the Config sheet.",
    "Edit here, commit the file, and the GitHub Action regenerates the site. No per-repo manifests needed.",
    "",
    "SHEETS",
    "1.  LISTINGS  = the cards on the site (libraries, SDKs, examples). One row per listing.",
    "2.  BOARDS    = the hardware registry. One row per dev board, with image filename + local fallback.",
    "3.  RESOURCES = per-board (or per-manufacturer) links: Buy, QuickStart, Developer guide, Demo,",
    "                Webinar, Video, Blog, Doc, Info. These power the board detail drawer on the site.",
    "4.  CONFIG    = key/value settings (org, image hosts, default status, pages URL).",
    "",
    "HOW IT FITS TOGETHER",
    "•  A listing points at one or more boards by Part Number (the Boards column).",
    "     - A 'sample' with several boards becomes several cards (one per board).",
    "     - A 'library' or 'sdk' with several boards stays one card listing them all.",
    "     - Leave Boards blank for platform-agnostic libraries/SDKs.",
    "•  Topics = comma-separated tags users filter by (ai, vision, audio, robotics, lora, …).",
    "•  RESOURCES rows join to a board by its Part Number in the Ref column (RefType=board),",
    "     or to a manufacturer by name (RefType=manufacturer). Kind drives the icon/label.",
    "•  BOARDS.Image File   = filename hosted at Config IMAGE_BASE (Azure).",
    "   BOARDS.Image Local  = filename committed under assets/boards/ (offline fallback).",
    "•  Set Include = no to hide a row from the site without deleting it.",
]
for line in INSTRUCTIONS:
    ins.append([line])
ins.column_dimensions["A"].width = 110
ins["A1"].font = Font(bold=True, size=13)

# Config
cfg_rows = [
    {"Key": "ORG", "Value": "avnet-iotconnect"},
    {"Key": "IMAGE_BASE", "Value": "https://saleshosted.z13.web.core.windows.net/images/boards/"},
    {"Key": "IMAGE_LOCAL_BASE", "Value": "assets/boards/"},
    {"Key": "BRAND_BASE", "Value": "assets/brand/"},
    {"Key": "DEFAULT_STATUS", "Value": "beta"},
    {"Key": "PAGES_URL", "Value": "https://mlamp99.github.io/iotc-index/"},
]
write_sheet("Config", ["Key", "Value"], cfg_rows, widths=[20, 70])
write_sheet("Listings", LCOLS, listings,
            widths=[34, 10, 10, 30, 14, 26, 34, 70, 46, 8])
write_sheet("Boards", BCOLS, list(boards.values()),
            widths=[16, 42, 26, 26, 26, 46, 14, 12, 22, 8])
write_sheet("Resources", ["Ref", "RefType", "Kind", "Title", "URL"], res_rows,
            widths=[26, 14, 12, 48, 70])

# order sheets: Instructions, Config, Listings, Boards, Resources
WB._sheets.sort(key=lambda s: ["Instructions", "Config", "Listings", "Boards", "Resources"].index(s.title))
WB.save(os.path.join(ROOT, "iotc-index-catalog.xlsx"))

print(f"listings : {len(listings)}  (+{added} new)")
print(f"boards   : {len(boards)}")
print(f"resources: {len(res_rows)}  (buy={sum(1 for r in res_rows if r['Kind']=='buy')}, "
      f"quickstart={sum(1 for r in res_rows if r['Kind']=='quickstart')}, "
      f"demo={sum(1 for r in res_rows if r['Kind']=='demo')}, "
      f"video={sum(1 for r in res_rows if r['Kind']=='video')})")
print(f"boards with local image: {sum(1 for b in boards.values() if b.get('Image Local'))}")
if orphan_res:
    print("WARN orphan resource refs:", orphan_res)
else:
    print("orphan resource refs: none")
