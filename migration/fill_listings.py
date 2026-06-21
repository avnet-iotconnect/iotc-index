#!/usr/bin/env python3
"""
fill_listings.py — one-time fill of missing Listings cells in the source workbook.

Fills Languages / Link / Repo / Topics from (a) GitHub primary languages,
(b) directly-queried non-org repo languages, and (c) web research of the
repo-less marketing demos (migration/_research.json). Also adds an "Image"
column so every board-less listing (SDKs, Drone, Telehealth) shows a relevant
picture (a language/tech logo or a real photo).

Edits ../iotc-index-catalog.xlsx in place (other sheets untouched).
"""
import os, re, json
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
XLSX = os.path.join(ROOT, "iotc-index-catalog.xlsx")

repoLang = json.load(open(os.path.join(HERE, "_listings.json"), encoding="utf-8"))["repoLang"]
patches  = json.load(open(os.path.join(HERE, "_research.json"), encoding="utf-8"))["patches"]

def norm(s):
    return re.sub(r"\s+", " ", str(s or "").replace("\xa0", " ").replace("&amp;", "&")).strip()

patch_by = {norm(p["name"]): p for p in patches}

# languages for non-org repos / repos GitHub didn't classify
SPECIAL_REPO_LANG = {
    "qcs6490-vision-ai-demo": "Python", "rasynboard-out-of-box-demo": "C", "spark": "C++",
    "stsafe-demo": "Python", "iotc-gateway-mobile-app": "Mobile",
    "iotc-python-greengrass-demos": "Python", "iotc-jetson-demo": "Python",
}

# board-less listings -> a relevant picture (tech logo or real photo)
IMAGE_MAP = {norm(k): v for k, v in {
    "/IOTCONNECT Core C Library": "assets/tech/c.svg",
    "/IOTCONNECT Python Library": "assets/tech/python.svg",
    "/IOTCONNECT Greengrass C Component SDK": "assets/tech/c.svg",
    "/IOTCONNECT Python Greengrass Components": "assets/tech/python.svg",
    "X-CUBE ST67 Wi-Fi SDK": "assets/tech/c.svg",
    "/IOTCONNECT Node.js SDK": "assets/tech/nodejs.svg",
    "/IOTCONNECT .NET SDK": "assets/tech/dotnet.svg",
    "/IOTCONNECT iOS Swift SDK": "assets/tech/swift.svg",
    "/IOTCONNECT Android SDK": "assets/tech/android.svg",
    "/IOTCONNECT Generic C SDK": "assets/tech/c.svg",
    "/IOTCONNECT Yocto C SDK": "assets/tech/linux.svg",
    "/IOTCONNECT Yocto Python SDK": "assets/tech/linux.svg",
    "/IOTCONNECT Python SDK": "assets/tech/python.svg",
    "/IOTCONNECT Python REST API": "assets/tech/python.svg",
    "Home Assistant Bridge": "assets/tech/homeassistant.svg",
    "IoTConnect Relay Service": "assets/tech/python.svg",
    "Arduino Uno Q App Lab": "assets/tech/arduino.svg",
    "Drone": "assets/listings/drone.png",
    "Telehealth Mobile Gateway": "assets/listings/telehealth.png",
}.items()}

wb = load_workbook(XLSX)
ws = wb["Listings"]
hdr = [c.value for c in ws[1]]
col = {h: i for i, h in enumerate(hdr)}

# ensure an "Image" column exists (append after Include)
if "Image" not in col:
    img_idx = len(hdr)
    ws.cell(row=1, column=img_idx + 1, value="Image")
    # match header styling of the neighbour
    ws.cell(row=1, column=img_idx + 1).font = ws.cell(row=1, column=1).font.copy()
    ws.cell(row=1, column=img_idx + 1).fill = ws.cell(row=1, column=1).fill.copy()
    col["Image"] = img_idx
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(img_idx + 1)].width = 26

def get(row, name):
    c = ws.cell(row=row, column=col[name] + 1); return c.value
def setc(row, name, val):
    ws.cell(row=row, column=col[name] + 1, value=val)

filled = {"Languages": 0, "Link": 0, "Repo": 0, "Topics": 0, "Image": 0, "Name": 0}
for r in range(2, ws.max_row + 1):
    name_raw = get(r, "Name")
    if not name_raw or not str(name_raw).strip():
        continue
    nm = norm(name_raw)
    repo = str(get(r, "Repo") or "").strip()
    p = patch_by.get(nm)

    # fix non-breaking space in the Name itself
    if "\xa0" in str(name_raw):
        setc(r, "Name", str(name_raw).replace("\xa0", " ")); filled["Name"] += 1

    # Languages
    if not str(get(r, "Languages") or "").strip():
        lang = None
        if repo and repoLang.get(repo.lower()):
            lang = repoLang[repo.lower()]
        elif repo and repo.lower() in SPECIAL_REPO_LANG:
            lang = SPECIAL_REPO_LANG[repo.lower()]
        elif p and p.get("languages"):
            lang = p["languages"]
        if lang:
            setc(r, "Languages", lang); filled["Languages"] += 1

    # Topics (only the one with a unicode-name mismatch was left empty)
    if not str(get(r, "Topics") or "").strip():
        if nm == norm("Image Classification and Remote Training"):
            setc(r, "Topics", "edge-ai, vision, mpu, ota"); filled["Topics"] += 1

    # Repo — only adopt a clean avnet-iotconnect org repo where it's the demo's real home
    if p and not repo and nm == norm("Telehealth Mobile Gateway"):
        setc(r, "Repo", "iotc-gateway-mobile-app"); repo = "iotc-gateway-mobile-app"; filled["Repo"] += 1

    # Link — adopt the researched best link (it already kept good existing ones);
    # otherwise derive from the repo.
    link = str(get(r, "Link") or "").strip()
    if p and p.get("link"):
        if p["link"] != link:
            setc(r, "Link", p["link"]); filled["Link"] += 1
    elif not link and repo:
        setc(r, "Link", f"https://github.com/avnet-iotconnect/{repo}"); filled["Link"] += 1

    # Image (board-less listings)
    if nm in IMAGE_MAP and not str(get(r, "Image") or "").strip():
        setc(r, "Image", IMAGE_MAP[nm]); filled["Image"] += 1

wb.save(XLSX)
print("filled:", filled)
